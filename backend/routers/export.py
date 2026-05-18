import io
import re
from urllib.parse import quote

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import Response
from sqlalchemy.orm import Session

from db.connection import get_db
from db import models

router = APIRouter(prefix="/api")


def _safe_filename(name: str, ext: str) -> str:
    ascii_name = re.sub(r"[^\w\s-]", "", name, flags=re.ASCII).strip()[:50] or "report"
    encoded = quote(name[:50].encode("utf-8"))
    return f"attachment; filename=\"{ascii_name}.{ext}\"; filename*=UTF-8''{encoded}.{ext}"


def _export_pdf(report: models.Report) -> Response:
    try:
        import markdown as md_lib
        from weasyprint import HTML
    except ImportError:
        raise HTTPException(status_code=501, detail="weasyprint 미설치. pip install weasyprint 후 brew install pango 필요.")

    body_html = md_lib.markdown(
        report.content,
        extensions=["tables", "fenced_code", "nl2br"],
    )
    html = f"""<!DOCTYPE html>
<html lang="ko">
<head>
<meta charset="utf-8"/>
<style>
  body {{
    font-family: -apple-system, 'Apple SD Gothic Neo', 'Noto Sans KR', 'Malgun Gothic', sans-serif;
    max-width: 720px; margin: 48px auto; padding: 0 24px;
    color: #1e293b; line-height: 1.85; font-size: 14px;
  }}
  .meta {{
    font-size: 11px; color: #64748b;
    padding: 10px 16px; background: #f8fafc;
    border-radius: 6px; margin-bottom: 2rem;
    display: flex; gap: 12px; align-items: center;
  }}
  h1 {{ font-size: 1.5rem; border-bottom: 2px solid #e2e8f0; padding-bottom: 0.4rem; margin-top: 2rem; }}
  h2 {{ font-size: 1.2rem; margin-top: 1.6rem; color: #1e40af; }}
  h3 {{ font-size: 1rem; margin-top: 1.2rem; }}
  p  {{ margin: 0.6rem 0; }}
  ul, ol {{ padding-left: 1.5rem; margin: 0.5rem 0; }}
  li {{ margin: 0.25rem 0; }}
  strong {{ font-weight: 700; }}
  code {{ background: #f1f5f9; padding: 0.1rem 0.3rem; border-radius: 4px; font-size: 0.85em; }}
  pre  {{ background: #f1f5f9; padding: 1rem; border-radius: 8px; overflow: auto; font-size: 0.85em; }}
  table {{ border-collapse: collapse; width: 100%; margin: 1rem 0; }}
  th, td {{ border: 1px solid #e2e8f0; padding: 6px 12px; text-align: left; }}
  th {{ background: #f8fafc; font-weight: 700; }}
</style>
</head>
<body>
  <div class="meta">
    <strong>Cosmic Hustle</strong>
    <span>{report.topic}</span>
  </div>
  {body_html}
</body>
</html>"""

    pdf_bytes = HTML(string=html).write_pdf()
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": _safe_filename(str(report.topic), "pdf")},
    )


def _export_excel(report: models.Report) -> Response:
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise HTTPException(status_code=501, detail="openpyxl 미설치. pip install openpyxl")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "리포트"
    ws.column_dimensions["A"].width = 110

    header_fill = PatternFill("solid", fgColor="1E3A5F")
    section_fill = PatternFill("solid", fgColor="0D1830")

    row = 1
    # 상단 제목
    cell = ws.cell(row=row, column=1, value=report.topic)
    cell.font = Font(bold=True, size=14, color="93C5FD")
    cell.fill = header_fill
    cell.alignment = Alignment(wrap_text=True)
    row += 2

    for line in report.content.splitlines():
        if line.startswith("# "):
            cell = ws.cell(row=row, column=1, value=line[2:])
            cell.font = Font(bold=True, size=13, color="93C5FD")
            cell.fill = header_fill
        elif line.startswith("## "):
            cell = ws.cell(row=row, column=1, value=line[3:])
            cell.font = Font(bold=True, size=12, color="C4B5FD")
            cell.fill = section_fill
        elif line.startswith("### "):
            cell = ws.cell(row=row, column=1, value=line[4:])
            cell.font = Font(bold=True, size=11, color="E2E8F0")
        elif line.startswith(("- ", "* ")):
            cell = ws.cell(row=row, column=1, value="  • " + line[2:])
            cell.font = Font(size=10, color="94A3B8")
        elif line.strip():
            cell = ws.cell(row=row, column=1, value=line.strip())
            cell.font = Font(size=10, color="CBD5E1")
            cell.alignment = Alignment(wrap_text=True)
        else:
            row += 1
            continue
        row += 1

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": _safe_filename(str(report.topic), "xlsx")},
    )


@router.get("/reports/{report_id}/export")
def export_report(
    report_id: str,
    format: str = Query(..., pattern="^(pdf|excel)$"),
    db: Session = Depends(get_db),
):
    report = db.query(models.Report).filter(models.Report.id == report_id).first()
    if not report:
        raise HTTPException(status_code=404, detail="Report not found")

    if format == "pdf":
        return _export_pdf(report)
    return _export_excel(report)
